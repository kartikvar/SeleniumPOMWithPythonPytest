import openpyxl


def total_row_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def total_column_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def read_data_from_excel(file, sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=colNum).value


def write_into_excel(file, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    (sheet.cell(row=rowNum, column=colNum).value) = data
    workbook.save(file)
